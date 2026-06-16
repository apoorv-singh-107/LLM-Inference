"""
AI Endpoint Benchmarking Script
Evaluates SSE streaming endpoint: POST /generative_response
Focuses on: brain/text and brain/body events
"""

import json
import time
import requests
import pandas as pd
from datetime import datetime
from dataclasses import dataclass
from typing import Optional
import traceback

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
ENDPOINT = "http://0.0.0.0:8000/generative_response"
EXCEL_PATH = "slm_exploration.xlsx"  # source questions
OUTPUT_EXCEL = "benchmark_results.xlsx"  # results file
USER_ID = "rak_ceramics"
ENABLE_THINKING = False
TEMPERATURE = 0.7
TOP_P = 0.95
TOP_K = 20
MIN_P = 0.0
PRESENCE_PENALTY = 1.5
WEB_SEARCH = True
REQUEST_TIMEOUT = 300  # seconds per request before hard abort
INTER_REQUEST_DELAY = 2  # seconds between requests (be kind to the server)

# Which questions to run (0-indexed). None = all
RUN_QUESTIONS: Optional[list] = None  # e.g. [0, 1, 5] to run only those rows


# ─────────────────────────────────────────────
#  DATA STRUCTURES
# ─────────────────────────────────────────────
@dataclass
class FunctionCall:
    """A single function/tool invocation captured from the stream."""

    agent: str  # "brain" | "querifai" | other sub-agent
    tool_name: str
    arguments: dict


@dataclass
class BrainBodyEvent:
    """Parsed brain/body event."""

    response: Optional[str]  # None on intermediate tool-dispatch events
    function_calls: list  # tool_invoke_args from brain
    prompt_tokens: int
    completion_tokens: int
    total_tokens: int
    raw_event_index: int  # position in stream


@dataclass
class BenchmarkResult:
    # ── Identity
    question_index: int
    question: str
    complexity: str

    # ── Timing
    request_sent_at: str  # ISO timestamp
    total_time_sec: float
    time_to_first_brain_token_sec: float
    brain_generation_start_sec: float  # when brain/text tokens first appeared
    brain_generation_end_sec: float  # when brain/text tokens stopped

    # ── Throughput
    brain_completion_tokens: int
    brain_prompt_tokens: int
    tokens_per_sec: float  # brain completion tokens / brain generation window

    # ── Agentic Steps
    brain_tool_calls: (
        list  # tools brain dispatched (e.g. retrieval_augmented_generation)
    )
    sub_agent_tool_calls: (
        list  # tools sub-agents used (e.g. get_tables, execute_mysql_query)
    )
    agentic_steps_summary: str  # human-readable ordered call trace

    # ── Response quality
    final_response: str  # assembled final answer
    response_char_count: int
    brain_body_count: int  # how many brain/body events fired
    stream_ended_cleanly: bool  # did we get a final brain/body with non-null response?
    premature_break_reason: str  # empty if clean

    # ── Error
    error: str = ""

    # ── Raw token counters per event source
    agents_total_tokens: int = 0

    # ── Extra observations
    notes: str = ""


# ─────────────────────────────────────────────
#  SSE PARSER
# ─────────────────────────────────────────────
def parse_sse_stream(response_iter):
    """
    Yields (event_type, data_dict, raw_line_index, recv_timestamp)
    from a raw SSE response iterator.
    """
    current_event = None
    current_data = None
    line_index = 0

    for raw_bytes in response_iter:
        recv_ts = time.perf_counter()
        line = raw_bytes.decode("utf-8", errors="replace").rstrip("\n").rstrip("\r")
        line_index += 1

        if line.startswith("event:"):
            current_event = line[len("event:") :].strip()
        elif line.startswith("data:"):
            current_data = line[len("data:") :].strip()
        elif line == "":
            # Blank line = end of one SSE message
            if current_event and current_data:
                try:
                    parsed = json.loads(current_data)
                except json.JSONDecodeError:
                    parsed = {"_raw": current_data}
                yield current_event, parsed, line_index, recv_ts
            current_event = None
            current_data = None


# ─────────────────────────────────────────────
#  SINGLE QUESTION RUNNER
# ─────────────────────────────────────────────
def run_question(idx: int, question: str, complexity: str) -> BenchmarkResult:
    print(f"\n{'=' * 70}")
    print(f"  Q{idx:02d}: {question[:80]}...")
    print(f"{'=' * 70}")

    payload = {
        "user_message": question,
        "sampling_paras": {
            "extra_body": {
                "chat_template_kwargs": {"enable_thinking": ENABLE_THINKING},
                "top_k": TOP_K,
                "min_p": MIN_P,
                "presence_penalty": PRESENCE_PENALTY,
            },
            "temperature": TEMPERATURE,
            "top_p": TOP_P,
        },
        "metadata": {
            "stream": True,
            "web_search": WEB_SEARCH,
            "user_id": USER_ID,
        },
        "chat_hist": None,
    }

    # Accumulators
    brain_token_chunks = []  # (text, recv_ts)
    brain_body_events: list[BrainBodyEvent] = []
    all_function_calls: list[FunctionCall] = []
    step_trace: list[str] = []  # ordered trace of agent actions
    agents_total_tokens = 0

    first_event_ts = None
    first_brain_text_ts = None
    brain_gen_start_ts = None
    brain_gen_end_ts = None
    last_event_ts = None
    stream_ended_cleanly = False
    premature_break_reason = ""
    error_msg = ""
    request_sent_at = datetime.now().isoformat()
    t_start = time.perf_counter()

    try:
        with requests.post(
            ENDPOINT,
            json=payload,
            headers={"Content-Type": "application/json"},
            stream=True,
            timeout=REQUEST_TIMEOUT,
        ) as resp:
            resp.raise_for_status()
            event_count = 0

            for event_type, data, line_idx, recv_ts in parse_sse_stream(
                resp.iter_lines()
            ):
                last_event_ts = recv_ts
                if first_event_ts is None:
                    first_event_ts = recv_ts

                event_count += 1

                # ── brain/body ──────────────────────────────────────────────
                if event_type == "brain/body":
                    response_text = data.get("response")
                    usage = data.get("metadata", {}).get("usage", {}) or {}
                    fc_raw = data.get("function_call") or {}
                    tool_invoke_args = fc_raw.get("tool_invoke_args", [])

                    bbe = BrainBodyEvent(
                        response=response_text,
                        function_calls=tool_invoke_args,
                        prompt_tokens=usage.get("prompt_tokens", 0),
                        completion_tokens=usage.get("completion_tokens", 0),
                        total_tokens=usage.get("total_tokens", 0),
                        raw_event_index=event_count,
                    )
                    brain_body_events.append(bbe)

                    for ti in tool_invoke_args:
                        name = ti.get("name", "unknown")
                        args = ti.get("arguments", {})
                        all_function_calls.append(FunctionCall("brain", name, args))
                        step_trace.append(f"[brain] → tool: {name}")
                        print(f"  🧠 Brain tool call: {name}")

                    if response_text is not None:
                        stream_ended_cleanly = True
                        print(
                            f"  ✅ Final brain/body received ({len(response_text)} chars)"
                        )

                # ── brain/text (streaming tokens from brain model) ──────────
                elif event_type == "brain/text":
                    parts = data.get("content", {}).get("parts", [])
                    for part in parts:
                        text = part.get("text", "")
                        if text:
                            brain_token_chunks.append((text, recv_ts))
                            if first_brain_text_ts is None:
                                first_brain_text_ts = recv_ts
                                brain_gen_start_ts = recv_ts
                            brain_gen_end_ts = recv_ts

                # ── sub-agent events (brain/querifai/text, etc.) ────────────
                elif "/" in event_type and event_type not in (
                    "brain/body",
                    "brain/text",
                ):
                    # Extract sub-agent name from event type
                    parts_path = event_type.split("/")
                    sub_agent = parts_path[1] if len(parts_path) > 1 else "unknown"

                    content = data.get("content", {})
                    if isinstance(content, dict):
                        parts = content.get("parts", [])
                    else:
                        parts = []

                    for part in parts:
                        # Sub-agent tool call
                        if "function_call" in part:
                            fc = part["function_call"]
                            tool_name = fc.get("name", "unknown")
                            tool_args = fc.get("args", {})
                            all_function_calls.append(
                                FunctionCall(sub_agent, tool_name, tool_args)
                            )
                            step_trace.append(f"[{sub_agent}] → tool: {tool_name}")
                            print(f"  🔧 {sub_agent} tool call: {tool_name}")
                        # Sub-agent function result
                        elif "function_response" in part:
                            fr = part["function_response"]
                            tool_name = fr.get("name", "unknown")
                            step_trace.append(f"[{sub_agent}] ← result: {tool_name}")

                    # Accumulate sub-agent tokens
                    usage = data.get("usage") or {}
                    if usage:
                        agents_total_tokens += usage.get("total_token_count", 0)

    except requests.exceptions.Timeout:
        error_msg = f"Request timed out after {REQUEST_TIMEOUT}s"
        premature_break_reason = error_msg
        print(f"  ❌ TIMEOUT: {error_msg}")
    except requests.exceptions.ConnectionError as e:
        error_msg = f"Connection error: {e}"
        premature_break_reason = error_msg
        print(f"  ❌ CONNECTION ERROR: {e}")
    except Exception as e:
        error_msg = traceback.format_exc()
        premature_break_reason = str(e)
        print(f"  ❌ ERROR: {e}")

    t_end = time.perf_counter() if last_event_ts is None else last_event_ts
    total_time = t_end - t_start

    # ── Compute metrics ──────────────────────────────────────────────────────
    time_to_first_brain_token = (
        (first_brain_text_ts - t_start) if first_brain_text_ts else -1
    )
    brain_gen_duration = (
        (brain_gen_end_ts - brain_gen_start_ts)
        if brain_gen_start_ts and brain_gen_end_ts
        else 0.0
    )

    final_brain_body = next(
        (e for e in reversed(brain_body_events) if e.response is not None), None
    )
    brain_completion_tokens = (
        final_brain_body.completion_tokens if final_brain_body else 0
    )
    brain_prompt_tokens = final_brain_body.prompt_tokens if final_brain_body else 0

    tokens_per_sec = (
        brain_completion_tokens / brain_gen_duration
        if brain_gen_duration > 0 and brain_completion_tokens > 0
        else 0.0
    )

    # Assemble final response: last brain/body with non-null response
    final_response = final_brain_body.response if final_brain_body else ""

    # If no final brain/body but we have brain/text tokens, assemble from those
    if not final_response and brain_token_chunks:
        final_response = "".join(t for t, _ in brain_token_chunks)

    if not stream_ended_cleanly and not premature_break_reason:
        premature_break_reason = "No final brain/body with response field received"

    # De-duplicate step trace (keep order, remove exact consecutive duplicates)
    deduped_trace = []
    for step in step_trace:
        if not deduped_trace or deduped_trace[-1] != step:
            deduped_trace.append(step)

    # Summarise brain tool calls
    brain_tools = [f.tool_name for f in all_function_calls if f.agent == "brain"]
    sub_tools_dedup = []
    seen = set()
    for f in all_function_calls:
        if f.agent != "brain":
            key = f"{f.agent}/{f.tool_name}"
            if key not in seen:
                sub_tools_dedup.append(f"{f.tool_name} ({f.agent})")
                seen.add(key)

    print(
        f"\n  ⏱  Total: {total_time:.2f}s | "
        f"First token: {time_to_first_brain_token:.2f}s | "
        f"TPS: {tokens_per_sec:.1f} | "
        f"Clean: {stream_ended_cleanly}"
    )

    return BenchmarkResult(
        question_index=idx,
        question=question,
        complexity=complexity,
        request_sent_at=request_sent_at,
        total_time_sec=round(total_time, 3),
        time_to_first_brain_token_sec=round(time_to_first_brain_token, 3),
        brain_generation_start_sec=round(
            brain_gen_start_ts - t_start if brain_gen_start_ts else -1, 3
        ),
        brain_generation_end_sec=round(
            brain_gen_end_ts - t_start if brain_gen_end_ts else -1, 3
        ),
        brain_completion_tokens=brain_completion_tokens,
        brain_prompt_tokens=brain_prompt_tokens,
        tokens_per_sec=round(tokens_per_sec, 2),
        brain_tool_calls=brain_tools,
        sub_agent_tool_calls=sub_tools_dedup,
        agentic_steps_summary=" → ".join(deduped_trace)
        if deduped_trace
        else "(direct answer)",
        final_response=final_response,
        response_char_count=len(final_response),
        brain_body_count=len(brain_body_events),
        stream_ended_cleanly=stream_ended_cleanly,
        premature_break_reason=premature_break_reason
        if not stream_ended_cleanly
        else "",
        error=error_msg,
        agents_total_tokens=agents_total_tokens,
    )


# ─────────────────────────────────────────────
#  EXCEL WRITER
# ─────────────────────────────────────────────
def save_results(results: list[BenchmarkResult], output_path: str):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Build rows
    rows = []
    for r in results:
        rows.append(
            {
                "Q#": r.question_index,
                "Question": r.question,
                "Complexity": r.complexity,
                "Run At": r.request_sent_at,
                # ── Timing ──
                "Total Time (s)": r.total_time_sec,
                "Time to First Token (s)": r.time_to_first_brain_token_sec,
                "Brain Gen Start (s)": r.brain_generation_start_sec,
                "Brain Gen End (s)": r.brain_generation_end_sec,
                # ── Throughput ──
                "Brain Completion Tokens": r.brain_completion_tokens,
                "Brain Prompt Tokens": r.brain_prompt_tokens,
                "Tokens / Sec": r.tokens_per_sec,
                "Agents Total Tokens": r.agents_total_tokens,
                # ── Agentic ──
                "Brain Tool Calls": ", ".join(r.brain_tool_calls)
                if r.brain_tool_calls
                else "(none)",
                "Sub-Agent Tool Calls": ", ".join(r.sub_agent_tool_calls)
                if r.sub_agent_tool_calls
                else "(none)",
                "Agentic Steps": r.agentic_steps_summary,
                "# brain/body Events": r.brain_body_count,
                # ── Stream Health ──
                "Stream Ended Cleanly": "YES" if r.stream_ended_cleanly else "NO",
                "Premature Break Reason": r.premature_break_reason,
                # ── Response ──
                "Response Char Count": r.response_char_count,
                "Final Response": r.final_response,
                # ── Error ──
                "Error": r.error,
            }
        )

    df = pd.DataFrame(rows)

    # Try to load existing workbook, else create new
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        wb = Workbook()

    sheet_name = "Benchmark Results"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── Styles ──
    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Arial", size=9)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write header
    headers = list(df.columns)
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(
                    col_name
                    in (
                        "Final Response",
                        "Agentic Steps",
                        "Question",
                        "Premature Break Reason",
                    )
                ),
            )
            cell.border = border
            # Highlight failed rows
            if col_name == "Stream Ended Cleanly" and val == "NO":
                cell.fill = PatternFill(
                    "solid", start_color="FF0000", end_color="FF0000"
                )
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    # Column widths
    col_widths = {
        "Q#": 5,
        "Question": 45,
        "Complexity": 16,
        "Run At": 22,
        "Total Time (s)": 14,
        "Time to First Token (s)": 18,
        "Brain Gen Start (s)": 16,
        "Brain Gen End (s)": 14,
        "Brain Completion Tokens": 20,
        "Brain Prompt Tokens": 18,
        "Tokens / Sec": 12,
        "Agents Total Tokens": 18,
        "Brain Tool Calls": 28,
        "Sub-Agent Tool Calls": 34,
        "Agentic Steps": 55,
        "# brain/body Events": 16,
        "Stream Ended Cleanly": 18,
        "Premature Break Reason": 30,
        "Response Char Count": 18,
        "Final Response": 60,
        "Error": 30,
    }
    for col_idx, col_name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(
            col_name, 15
        )

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # ── Summary sheet ──
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum["A1"] = "Benchmark Summary"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14)

    summary_rows = [
        ("Total Questions Run", len(results)),
        ("Clean Streams", sum(1 for r in results if r.stream_ended_cleanly)),
        ("Premature Breaks", sum(1 for r in results if not r.stream_ended_cleanly)),
        (
            "Avg Total Time (s)",
            round(sum(r.total_time_sec for r in results) / len(results), 2)
            if results
            else 0,
        ),
        (
            "Avg First Token Time (s)",
            round(
                sum(
                    r.time_to_first_brain_token_sec
                    for r in results
                    if r.time_to_first_brain_token_sec > 0
                )
                / max(
                    1, sum(1 for r in results if r.time_to_first_brain_token_sec > 0)
                ),
                2,
            ),
        ),
        (
            "Avg Tokens/Sec",
            round(
                sum(r.tokens_per_sec for r in results if r.tokens_per_sec > 0)
                / max(1, sum(1 for r in results if r.tokens_per_sec > 0)),
                2,
            ),
        ),
        (
            "Total Brain Tokens Generated",
            sum(r.brain_completion_tokens for r in results),
        ),
        ("Run Timestamp", datetime.now().isoformat()),
    ]
    for i, (label, value) in enumerate(summary_rows, 3):
        ws_sum.cell(row=i, column=1, value=label).font = Font(
            name="Arial", bold=True, size=10
        )
        ws_sum.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10)
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 24

    wb.save(output_path)
    print(f"\n✅ Results saved → {output_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print("Loading questions from Excel...")
    df = pd.read_excel(EXCEL_PATH, sheet_name="Sheet1")

    questions = df["Question"].tolist()
    complexities = df["Complexity"].fillna("Unknown").tolist()

    indices = (
        RUN_QUESTIONS if RUN_QUESTIONS is not None else list(range(len(questions)))
    )

    print(f"Will run {len(indices)} question(s) against {ENDPOINT}")

    results: list[BenchmarkResult] = []

    for i, q_idx in enumerate(indices):
        question = str(questions[q_idx])
        complexity = str(complexities[q_idx])

        print(f"\n[{i + 1}/{len(indices)}] Question index {q_idx}")
        result = run_question(q_idx, question, complexity)
        results.append(result)

        # Save after each question so partial results aren't lost
        save_results(results, OUTPUT_EXCEL)

        if i < len(indices) - 1:
            print(f"  Waiting {INTER_REQUEST_DELAY}s before next request...")
            time.sleep(INTER_REQUEST_DELAY)

    print(f"\n{'=' * 70}")
    print(f"  BENCHMARK COMPLETE: {len(results)} questions")
    print(f"  Results: {OUTPUT_EXCEL}")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    main()
